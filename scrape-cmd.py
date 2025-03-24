import argparse
import xml.etree.ElementTree as ET
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def round_to_sig_figs(value, sig_figs=3):
    """Rounds a number to a specified number of significant figures."""
    try:
        num = float(value)
        if num == 0:
            return "0"  
        return f"{num:.{sig_figs}g}" 
    except ValueError:
        return value


def parse_xml(file_path):
    """Parses an XML or PDBXML file and extracts structured data for each battery cell."""
    tree = ET.parse(file_path)
    root = tree.getroot()

    # Extract general test details
    general_info = {}
    stringname = {}
    jarcells = {}
    deviation = {}
    tablesummary = {}

    for form in root.findall("form"):
        for test in form.findall("test"):
            general_info["Test Date"] = test.get("date")

            for data in test.findall("data"):
                temp_tag = data.find("tag[@name='temperature']")
                form_tag = data.find("tag[@name='formname']")
                if temp_tag is not None:
                    general_info["Ambient Temp."] = temp_tag.text  
                if form_tag is not None:
                    general_info["Form Name"] = form_tag.text

                avgimpedence_tag = data.find("tag[@name='avgimpedence']")
                totalstringvoltage_tag = data.find("tag[@name='voltagesum']")
                totaldeviationvolage_tag = data.find("tag[@name='deviationvoltage']")
                minvolt_tag = data.find("tag[@name='minvolts']")
                maxvolt_tag = data.find("tag[@name='maxvolts']")
                avgtemp_tag = data.find("tag[@name='avgtemp']")
                if avgimpedence_tag is not None:
                    tablesummary["Average Impedance (mΩ)"] = avgimpedence_tag.text
                if totalstringvoltage_tag is not None:
                    tablesummary["Total String Voltage (V)"] = totalstringvoltage_tag.text
                if totaldeviationvolage_tag is not None:
                    tablesummary["Deviation from Charger Voltage (%)"] = totaldeviationvolage_tag.text
                if minvolt_tag is not None:
                    tablesummary["Min Voltage (V)"] = minvolt_tag.text
                if maxvolt_tag is not None:
                    tablesummary["Max Voltage (V)"] = maxvolt_tag.text
                if avgtemp_tag is not None:
                    tablesummary["Average Temperature (°C)"] = avgtemp_tag.text
                
                

            for nameplate in test.findall("nameplate"):
                stringname_tag = nameplate.find("tag[@name='stringname']")
                equipmenttype_tag = nameplate.find("tag[@name='pdbequipmenttype']")
                
                if stringname_tag is not None:
                    stringname["String Name"] = stringname_tag.text
                if equipmenttype_tag is not None:
                    stringname["Battery Type"] = equipmenttype_tag.text

                deviationwarningohm_tag = nameplate.find("tag[@name='warningdeviationohm']")
                deviationalarmohm_tag = nameplate.find("tag[@name='alloweddeviationohm']")
                deviationwarning_tag = nameplate.find("tag[@name='warningdeviation']")
                deviationalarm_tag = nameplate.find("tag[@name='alloweddeviation']")
                if deviationwarningohm_tag is not None:
                    deviation["Warning Deviation (mΩ)"] = deviationwarningohm_tag.text
                if deviationalarmohm_tag is not None:
                    deviation["Alarm Deviation (mΩ)"] = deviationalarmohm_tag.text
                if deviationwarning_tag is not None:
                    deviation["Warning Deviation (%)"] = deviationwarning_tag.text
                if deviationalarm_tag is not None:
                    deviation["Alarm Deviation (%)"] = deviationalarm_tag.text

            for copyhistory in test.findall("copyhistory"):
                numjars_tag = copyhistory.find("tag[@name='numjars']")
                numcells_tag = copyhistory.find("tag[@name='numcells']")
                cellsperjar_tag = copyhistory.find("tag[@name='cellsperjar']")
                numstraps_tag = copyhistory.find("tag[@name='numstraps']")
                
                if numjars_tag is not None:
                    jarcells["Number of Jars"] = numjars_tag.text
                if numcells_tag is not None:
                    jarcells["Number of Cells"] = numcells_tag.text
                if cellsperjar_tag is not None:
                    jarcells["Number of Cells/Jar"] = cellsperjar_tag.text
                if numstraps_tag is not None:
                    jarcells["Number of Straps"] = numstraps_tag.text

                baseline_tag = copyhistory.find("tag[@name='baselinez']")
                if baseline_tag is not None:
                    baseline = baseline_tag.text


            # Extract per-cell data (organized horizontally)
            cell_data = []
            for array in test.findall(".//array"):
                array_name = array.get("name")

                # Process array values grouped by cell number
                for item in array.findall("arrayitem"):
                    cell_no = int(item.get("index"))
                    value = item.text if item.text is not None else ""

                    # Find or create entry for this cell
                    cell_entry = next((cell for cell in cell_data if cell["Cell No"] == cell_no), None)
                    if not cell_entry:
                        cell_entry = {"Cell No": cell_no}
                        cell_data.append(cell_entry)

                    cell_entry[array_name] = value
        
            return general_info, cell_data, stringname, jarcells, deviation, tablesummary, baseline

def write_excel(general_info, cell_data, stringname, jarcells, deviation, tablesummary, baseline, output_file):
    """Writes extracted data into a well-structured Excel (.xlsx) file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Battery Test"

    # Formatting Headers
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center")

    ws["A1"] = general_info.pop("Form Name", "Battery Test Report")
    ws["A1"].font = bold_font
    ws.append([])                   
    
    for i, (key, value) in enumerate(general_info.items()):
        key_cell = ws.cell(row=3 + i, column=1, value=key)
        key_cell.font = bold_font
        ws.cell(row=3 + i, column=2, value=value)    
    for i, (key, value) in enumerate(stringname.items()):
        key_cell = ws.cell(row=5 + i, column=1, value=key)
        key_cell.font = bold_font
        ws.cell(row=5 + i, column=2, value=value)
    for i, (key, value) in enumerate(jarcells.items()):
        key_cell = ws.cell(row=3 + i, column=3, value=key)
        key_cell.font = bold_font
        ws.cell(row=3 + i, column=4, value=value)

    ws.append([])

    current_row = ws.max_row + 2

    for i, (key, value) in enumerate(deviation.items()):
        value = round_to_sig_figs(value,5)
        if i < 2:
            key_cell = ws.cell(row=current_row + i, column=1, value=key)
            key_cell.font = bold_font
            ws.cell(row=current_row + i, column=2, value=value)
        else:
            key_cell = ws.cell(row=current_row + i -2, column=3, value=key)
            key_cell.font = bold_font
            ws.cell(row=current_row + i-2, column=4, value=value)
            
    ws.append([])
    ws.append(["Table Summary"])
    ws["A{}".format(ws.max_row)].font = bold_font
    current_row = ws.max_row + 1
    key_cell = ws.cell(row=current_row, column=1, value="Baseline Impedance (mΩ)")
    key_cell.font = bold_font
    ws.cell(row=current_row +1, column=1, value=baseline)
    
    for i, (key, value) in enumerate(tablesummary.items()):
        key_cell = ws.cell(row=current_row, column=2 + i, value=key)
        key_cell.font = bold_font
        value = round_to_sig_figs(value,5)
        ws.cell(row=current_row + 1, column=2 + i, value=value)


    ws.append([])  # Add empty row for spacing

    headers = ["Cell No.", "Impedance (mΩ)", "% Deviation (Baseline)", "% Variation (String)", "Voltage (V)", "Time", "Temperature (°C)"]
    ws.append(headers)

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws["{}{}".format(col, ws.max_row)].font = bold_font
        ws["{}{}".format(col, ws.max_row)].alignment = center_align

    for row in cell_data:
        ws.append([
            row.get("Cell No", ""),
            row.get("impedence", ""),
            row.get("v", ""),
            row.get("d", ""),
            row.get("voltage", ""),
            row.get("time", ""),
            row.get("tem", "")
        ])

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Save workbook
    wb.save(output_file)

def main():
    parser = argparse.ArgumentParser(description="Convert PDBXML/XML to Excel (.xlsx) with structured formatting")
    parser.add_argument("input_file", help="Path to the XML/PDBXML file")
    parser.add_argument("-o", "--output", help="Output Excel file name (default: input_file.xlsx)", default=None)

    args = parser.parse_args()
    input_file = args.input_file
    output_file = args.output or os.path.splitext(input_file)[0] + "_report.xlsx"

    if not os.path.exists(input_file):
        print(f"Error: File '{input_file}' not found.")
        return

    print(f"Processing: {input_file}")
    general_info, cell_data, stringname, jarcells, deviation, tablesummary, baseline = parse_xml(input_file)
    write_excel(general_info, cell_data, stringname, jarcells, deviation, tablesummary, baseline, output_file)
    
    print(f"✅ Conversion complete! Excel file saved as: {output_file}")

if __name__ == "__main__":
    main()