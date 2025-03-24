import os
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import Tk, filedialog, messagebox, Label, Button, Entry
from openpyxl import Workbook 
from openpyxl.styles import Font, Alignment
from openpyxl.chart import LineChart, Reference

def convert_to_number(value):
    """Converts a string to a number (int or float) if possible, otherwise returns the original string."""
    try:
        if "." in value: 
            return float(value)
        return int(value)  
    except (ValueError, TypeError):
        return value 

def round_to_sig_figs(value, sig_figs=3):
    """Rounds a number to a specified number of significant figures."""
    try:
        num = float(value)
        if num == 0:
            return "0"  
        return f"{num:.{sig_figs}g}" 
    except ValueError:
        return value


def parse_xml(file_path):
    """Parses an XML or PDBXML file and extracts structured data for each battery cell."""
    tree = ET.parse(file_path)
    root = tree.getroot()

    all_tests = []
    formname = ""
    for form in root.findall("form"):
        for test in form.findall("test"):
            general_info = {}
            stringname = {}
            jarcells = {}
            deviation = {}
            tablesummary = {}
            baseline = "N/A"

            general_info["Test Date"] = test.get("date")

            for data in test.findall("data"):
                temp_tag = next(
                    (tag for tag in data.findall("tag") if tag.get("name", "").lower() == "temperature"),
                    None
                )
                form_tag = next(
                    (tag for tag in data.findall("tag") if tag.get("name", "").lower() == "formname"),
                    None
                )

                if temp_tag is not None:
                    general_info["Ambient Temp. (°C)"] = temp_tag.text  
                if form_tag is not None:
                    formname = form_tag.text

                avgimpedence_tag = next(
                    (tag for tag in data.findall("tag") if tag.get("name", "").lower() == "avgimpedence"),
                    None
                )
                totalstringvoltage_tag = next(
                    (tag for tag in data.findall("tag") if tag.get("name", "").lower() == "voltagesum"),
                    None
                )
                totaldeviationvolage_tag = next(
                    (tag for tag in data.findall("tag") if tag.get("name", "").lower() == "deviationvoltage"),
                    None
                )
                minvolt_tag = next(
                    (tag for tag in data.findall("tag") if tag.get("name", "").lower() == "minvolts"),
                    None
                )
                maxvolt_tag = next(
                    (tag for tag in data.findall("tag") if tag.get("name", "").lower() == "maxvolts"),
                    None
                )
                avgtemp_tag = next(
                    (tag for tag in data.findall("tag") if tag.get("name", "").lower() == "avgtemp"),
                    None
                )
                if avgimpedence_tag is not None:
                    tablesummary["Average Impedance (mΩ)"] = avgimpedence_tag.text
                if totalstringvoltage_tag is not None:
                    tablesummary["Total String Voltage (V)"] = totalstringvoltage_tag.text
                if totaldeviationvolage_tag is not None:
                    tablesummary["Deviation from Charger Voltage (%)"] = totaldeviationvolage_tag.text
                if minvolt_tag is not None:
                    tablesummary["Min Voltage (V)"] = minvolt_tag.text
                if maxvolt_tag is not None:
                    tablesummary["Max Voltage (V)"] = maxvolt_tag.text
                if avgtemp_tag is not None:
                    tablesummary["Average Temperature (°C)"] = avgtemp_tag.text
                
                

            for nameplate in test.findall("nameplate"):
                stringname_tag = next(
                    (tag for tag in nameplate.findall("tag") if tag.get("name", "").lower() == "stringname"),
                    None
                )
                equipmenttype_tag = next(
                    (tag for tag in nameplate.findall("tag") if tag.get("name", "").lower() == "pdbequipmenttype"),
                    None
                )
                
                if stringname_tag is not None:
                    stringname["String Name"] = stringname_tag.text
                if equipmenttype_tag is not None:
                    stringname["Battery Type"] = equipmenttype_tag.text

                deviationwarningohm_tag = next(
                    (tag for tag in nameplate.findall("tag") if tag.get("name", "").lower() == "warningdeviationohm"),
                    None
                )
                deviationalarmohm_tag = next(
                    (tag for tag in nameplate.findall("tag") if tag.get("name", "").lower() == "alloweddeviationohm"),
                    None
                )
                deviationwarning_tag = next(
                    (tag for tag in nameplate.findall("tag") if tag.get("name", "").lower() == "warningdeviation"),
                    None
                )
                deviationalarm_tag = next(
                    (tag for tag in nameplate.findall("tag") if tag.get("name", "").lower() == "alloweddeviation"),
                    None
                )
                if deviationwarningohm_tag is not None:
                    deviation["Warning Deviation (mΩ)"] = deviationwarningohm_tag.text
                if deviationalarmohm_tag is not None:
                    deviation["Alarm Deviation (mΩ)"] = deviationalarmohm_tag.text
                if deviationwarning_tag is not None:
                    deviation["Warning Deviation (%)"] = deviationwarning_tag.text
                if deviationalarm_tag is not None:
                    deviation["Alarm Deviation (%)"] = deviationalarm_tag.text

            for copyhistory in test.findall("copyhistory"):
                numjars_tag = next(
                    (tag for tag in copyhistory.findall("tag") if tag.get("name", "").lower() == "numjars"),
                    None
                )
                numcells_tag = next(
                    (tag for tag in copyhistory.findall("tag") if tag.get("name", "").lower() == "numcells"),
                    None
                )
                cellsperjar_tag = next(
                    (tag for tag in copyhistory.findall("tag") if tag.get("name", "").lower() == "cellsperjar"),
                    None
                )
                numstraps_tag = next(
                    (tag for tag in copyhistory.findall("tag") if tag.get("name", "").lower() == "numstraps"),
                    None
                )
                
                if numjars_tag is not None:
                    jarcells["Number of Jars"] = numjars_tag.text
                if numcells_tag is not None:
                    jarcells["Number of Cells"] = numcells_tag.text
                if cellsperjar_tag is not None:
                    jarcells["Number of Cells/Jar"] = cellsperjar_tag.text
                if numstraps_tag is not None:
                    jarcells["Number of Straps"] = numstraps_tag.text

                baseline_tag = next(
                    (tag for tag in copyhistory.findall("tag") if tag.get("name", "").lower() == "instrbaselinez"),
                    None
                )
                if baseline_tag is not None:
                    baseline = baseline_tag.text

            cell_data = []
            for array in test.findall(".//array"):
                array_name = array.get("name")

                for item in array.findall("arrayitem"):
                    cell_no = int(item.get("index"))
                    value = item.text if item.text is not None else ""

                    cell_entry = next((cell for cell in cell_data if cell["Cell No"] == cell_no), None)
                    if not cell_entry:
                        cell_entry = {"Cell No": cell_no}
                        cell_data.append(cell_entry)

                    cell_entry[array_name] = value
        
            all_tests.append((general_info, cell_data, stringname, jarcells, deviation, tablesummary, baseline))
        return formname, all_tests

def write_excel(formname, all_tests, graph_bool, output_file):
    """Writes extracted data into a well-structured Excel (.xlsx) file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Battery Test"

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    if formname != "":
        ws["A1"] = formname
    else:
        ws["A1"] = "Battery Test Report"
    
    ws["A1"].font = bold_font
    ws.append([])                   
    current_row = ws.max_row
    for general_info, cell_data, stringname, jarcells, deviation, tablesummary, baseline in all_tests:
        current_row += 2
        for i, (key, value) in enumerate(general_info.items()):
            key_cell = ws.cell(row=current_row + i, column=1, value=key)
            key_cell.font = bold_font
            if key == "Ambient Temp. (°C)":
                value = convert_to_number(value)
            ws.cell(row=current_row + i, column=2, value=value)    
        for i, (key, value) in enumerate(stringname.items()):
            key_cell = ws.cell(row=current_row + 2 + i, column=1, value=key)
            key_cell.font = bold_font
            ws.cell(row=current_row + 2 + i, column=2, value=value)
        for i, (key, value) in enumerate(jarcells.items()):
            key_cell = ws.cell(row=current_row + i, column=3, value=key)
            key_cell.font = bold_font
            value = convert_to_number(value)
            ws.cell(row=current_row + i, column=4, value=value)

        ws.append([])

        current_row = ws.max_row + 2

        for i, (key, value) in enumerate(deviation.items()):
            value = round_to_sig_figs(value,5)
            value = convert_to_number(value)
            if i < 2:
                key_cell = ws.cell(row=current_row + i, column=1, value=key)
                key_cell.font = bold_font
                ws.cell(row=current_row + i, column=2, value=value)
            else:
                key_cell = ws.cell(row=current_row + i -2, column=3, value=key)
                key_cell.font = bold_font
                value = convert_to_number(value)
                ws.cell(row=current_row + i-2, column=4, value=value)
                
        ws.append([])
        ws.append(["Table Summary"])
        ws["A{}".format(ws.max_row)].font = bold_font
        current_row = ws.max_row + 1
        key_cell = ws.cell(row=current_row, column=1, value="Baseline Impedance (mΩ)")
        key_cell.font = bold_font
        baseline = convert_to_number(baseline)
        ws.cell(row=current_row +1, column=1, value=baseline)
        
        for i, (key, value) in enumerate(tablesummary.items()):
            key_cell = ws.cell(row=current_row, column=2 + i, value=key)
            key_cell.font = bold_font
            key_cell.alignment = center_align
            value = round_to_sig_figs(value,5)
            value = convert_to_number(value)
            ws.cell(row=current_row + 1, column=2 + i, value=value)


        ws.append([])

        headers = ["Cell No.", "Impedance (mΩ)", "% Deviation (Baseline)", "% Variation (String)", "Voltage (V)", "Time", "Temperature (°C)"]
        ws.append(headers)

        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws["{}{}".format(col, ws.max_row)].font = bold_font
            ws["{}{}".format(col, ws.max_row)].alignment = center_align

        for row in cell_data:
            row = {k.lower(): v for k, v in row.items()}
            ws.append([
                convert_to_number(row.get("cell no", "")),
                convert_to_number(row.get("impedence", "")),
                convert_to_number(row.get("v", "")),
                convert_to_number(row.get("d", "")),
                convert_to_number(row.get("voltage", "")),
                row.get("time", ""),
                convert_to_number(row.get("tem_1", ""))
            ])

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        if graph_bool:
            voltageChart = LineChart()
            voltageChart.title = "Voltage Graph"
            voltageChart.x_axis.title = "Cell Number"
            voltageChart.y_axis.title = "Voltage (V)"
            voltageChart.legend = None
            impedanceChart = LineChart()
            impedanceChart.title = "Impedance Graph"
            impedanceChart.x_axis.title = "Cell Number"
            impedanceChart.y_axis.title = "Impedance (mΩ)"
            impedanceChart.legend = None

            start_row = ws.max_row - len(cell_data) + 1  
            end_row = ws.max_row
            graph_row = start_row
            graph_col1 = "J"
            graph_col2 = "T"

            x_values = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)  
            yVoltage_values = Reference(ws, min_col=5, min_row=start_row, max_row=end_row)  
            yImpedance_values = Reference(ws, min_col=2, min_row=start_row, max_row=end_row) 

            voltageChart.add_data(yVoltage_values, titles_from_data=True)
            voltageChart.set_categories(x_values)
            impedanceChart.add_data(yImpedance_values, titles_from_data=True)
            impedanceChart.set_categories(x_values)

            ws.add_chart(impedanceChart, f"{graph_col1}{graph_row}")
            ws.add_chart(voltageChart, f"{graph_col2}{graph_row}")
        current_row = ws.max_row
    wb.save(output_file)

def select_files():
    """Open file dialog to select multiple XML/PDBXML files."""
    file_paths = filedialog.askopenfilenames(filetypes=[("PDBXML/XML Files", "*.pdbxml;*.xml")])
    if file_paths:
        file_entry.delete(0, 'end')
        file_entry.insert(0, "; ".join(file_paths))

def select_output_folder():
    """Open dialog to select output folder."""
    folder_path = filedialog.askdirectory()
    if folder_path:
        output_entry.delete(0, 'end')
        output_entry.insert(0, folder_path)

def convert_files():
    """Handles the conversion of one or more XML/PDBXML files to Excel."""
    input_files = file_entry.get().split("; ")
    output_folder = output_entry.get()
    graph_bool = graph_var.get()
    
    if not input_files or not os.path.exists(input_files[0]):
        messagebox.showerror("Error", "Please select valid XML/PDBXML files.")
        return
    
    if not output_folder or not os.path.exists(output_folder):
        output_folder = os.path.dirname(input_files[0])
        output_entry.insert(0, output_folder)
    
    try:
        for input_file in input_files:
            filename = os.path.basename(input_file).replace(".xml", "").replace(".pdbxml", "")
            output_file = os.path.join(output_folder, f"{filename}_report.xlsx")
            formname, all_tests = parse_xml(input_file)
            write_excel(formname, all_tests, graph_bool, output_file)
        
        messagebox.showinfo("Success", f"Conversion complete! Excel files saved in {output_folder}.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{e}")

root = Tk()
root.title("PDBXML to Excel Converter (Multiple Files)")
root.geometry("600x300")

Label(root, text="Select XML/PDBXML Files:").pack(pady=5)
file_entry = Entry(root, width=70)
file_entry.pack()
Button(root, text="Browse", command=select_files).pack()

graph_var = tk.BooleanVar(value=True)
tk.Checkbutton(root, text="Auto Generate Graphs", variable=graph_var).pack()

Label(root, text="Select Output Folder: (default location of first file)").pack(pady=5)
output_entry = Entry(root, width=70)
output_entry.pack()
Button(root, text="Browse", command=select_output_folder).pack()

Button(root, text="Convert", command=convert_files, fg="white", bg="green").pack(pady=20)

root.mainloop()